#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Created on Fri Jan 15 12:16:59 2021

@author: emilyfang
"""

import pandas as pd
import openpyxl

# #CALCULATIONS FUNCTION

def calculations(df):
    
    #Population of Color
    df['Population of Color (%)'] = (df['Total with Race Data']-df['Total White Alone']) / df['Total with Race Data'] * 100
    #Neighborhood Renters who are Cost-Burdened
    df['% Neighborhood Renters who are Cost-Burdened'] = (df['Rent 30.0-34.9%'] \
                                                          + df['Rent 35.0-39.9%']
                                                          + df['Rent 40.0-49.9%']
                                                          + df['Rent >50.0%']) / df['Total with Rent Data'] * 100
    #Educational Attainment
    df['% with Associates/Bachelors Degree or higher'] = df['% >25 with Associates'] \
                                                            + df['% >25 with Bachelors or higher']
    return df

def touchups_and_export(df):
    #REMOVE COLUMNS
    remove_list = ['Total with Race Data', 'Total White Alone', \
                    'Total with Rent Data', 'Rent 30.0-34.9%', 'Rent 35.0-39.9%', 'Rent 35.0-39.9%', 'Rent 40.0-49.9%', 'Rent >50.0%']
    df.drop(remove_list, axis=1, inplace=True)
    
    #METRIC ORDER
    df_ORHD = df[['Life Expectancy', \
                  'Cancer (excluding skin cancer) among adults >= 18', \
                  'Current asthma among adults >= 18', \
                  'COPD among adults >= 18', \
                  'Coronary heart disease among adults >= 18', \
                  'Diabetes among adults >= 18', \
                  'Stroke among adults >= 18', \
                  'Mental health not good for >= 14 days among adults >= 18', \
                  'PM 2.5 (ug/m3)']]
    #df_ORHD = df[[]]
    df_TAU = df[['% Public Transit', '% Walked', '% Bicycle']]
    df_OARE = df[['Median Household Income', \
              'Poverty Rate', \
              'Unemployment Rate', \
              '% with Associates/Bachelors Degree or higher', \
              'Population of Color (%)', '% Limited English Households', \
              '% Neighborhood Renters who are Cost-Burdened']]
    
    df_ORHD = df_ORHD.T  
    df_TAU = df_TAU.T
    df_OARE = df_OARE.T
    
    #EXCEL SHEET
    sheet_title = 'HealthScore 2.0'
    writer = pd.ExcelWriter('test.xlsx',engine='xlsxwriter')   
    workbook=writer.book
    worksheet=workbook.add_worksheet(sheet_title)
    writer.sheets[sheet_title] = worksheet
    df_ORHD.to_excel(writer,sheet_name=sheet_title,startrow=1 , startcol=0)   
    df_TAU.to_excel(writer,sheet_name=sheet_title,startrow=len(df_ORHD)+4+2, startcol=0)
    df_OARE.to_excel(writer,sheet_name=sheet_title,startrow=len(df_ORHD)+4 + len(df_TAU)+4+3, startcol=0)
    writer.save()
    
    
    srcfile = openpyxl.load_workbook('test.xlsx',read_only=False, keep_vba= True)
    sheetname = srcfile.get_sheet_by_name(sheet_title)
    
    sheetname.cell(row=1,column=1).value = "Opportunity to Reduce Health Disparities" 
    sheetname.cell(row=len(df_ORHD)+4+2,column=1).value = "Transportation Access and Utilization" 
    sheetname.cell(row=len(df_ORHD)+4 + len(df_TAU)+4+3,column=1).value = "Opportunity to Advance Regional Equity" 
    
    
    srcfile.save('test.xlsm')
    
    print('DONE')
    



